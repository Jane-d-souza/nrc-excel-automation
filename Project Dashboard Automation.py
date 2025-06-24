from openpyxl import Workbook, load_workbook
from copy import copy
import os 

#-----------COPY FUNCTIONS--------------
def copy_sheet(source_ws, target_ws):
    copy_cells(source_ws, target_ws)
    copy_sheet_attributes(source_ws, target_ws)

def copy_sheet_attributes(source_ws, target_ws):
    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.merged_cells = copy(source_ws.merged_cells)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.freeze_panes = copy(source_ws.freeze_panes)

    for rn in range(len(source_ws.row_dimensions)):
        target_ws.row_dimensions[rn] = copy(source_ws.row_dimensions[rn])

    if source_ws.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_ws.sheet_format.defaultColWidth = copy(source_ws.sheet_format.defaultColWidth)

    for key, value in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].min = copy(value.min)
        target_ws.column_dimensions[key].max = copy(value.max)
        target_ws.column_dimensions[key].width = copy(value.width)
        target_ws.column_dimensions[key].hidden = copy(value.hidden)

def copy_cells(source_ws, target_ws):
    for (row, col), source_cell in source_ws._cells.items():
        target_cell = target_ws.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

#-----------END OF COPY FUNCTIONS--------------



# --- Main Script ---
# Financial Source Workbook
financial_wb = load_workbook(r'C:\Users\Dsouzaj\Documents\NRC Python Files\ATAC NRC Financial Report Mar 2025.xlsx', data_only=True)
financial_ws = financial_wb['USD Monthly Totals']

# Dashboard Template Workbook
source_wb = load_workbook(r'C:\Users\Dsouzaj\Documents\NRC Python Files\ATAC Project Dashbaord_Feb 28, 25_v1.xlsx')
sheet_names = ['Presentation Working Sheet', 'NRC ATAC All Phases', 'CAD to USD Savings']

# Target Workbook
target_wb = Workbook()
if 'Sheet' in target_wb.sheetnames:
    std = target_wb['Sheet']
    target_wb.remove(std)

# Loop through each sheet and copy
for name in sheet_names:
    source_ws = source_wb[name]
    target_ws = target_wb.create_sheet(title=name)
    copy_sheet(source_ws, target_ws)

# --- Custom cell logic for "Presentation Working Sheet" ALL PHASES TABLE-----
presentation_ws = target_wb['Presentation Working Sheet']

# Get original values
g9 = presentation_ws['G9'].value or 0
h9 = presentation_ws['H9'].value or 0

# Sum G9 and H9, put result in G9
presentation_ws['G9'].value = g9 + h9

# Replace H9 with value from AJ31 in financial report
aj31_value = financial_ws['AJ31'].value
presentation_ws['H9'].value = aj31_value

# --- Custom cell logic for "Presentation Working Sheet" LABOUR + TRAVEL-----

g22 = presentation_ws['G22'].value

# Get value from AI33 in the financial report
ai33_value =  financial_ws['AI33'].value

# Sum and put result in G22
presentation_ws['G22'].value = g22 + ai33_value

# Sum AJ17 and AJ21 from the financial report and put in H22
aj17_value =  financial_ws['AJ17'].value
aj21_value =  financial_ws['AJ21'].value
presentation_ws['H22'].value = aj17_value + aj21_value



# --- Custom cell logic for "Presentation Working Sheet" LABOUR + TRAVEL ACTUALS-----

presentation_ws['H40'].value = aj17_value
presentation_ws['H41'].value = aj21_value


# --- Custom cell logic for "Presentation Working Sheet" NRE + SI -----

# Sum G53 and H53 in Presentation and put result in G53
g53 = presentation_ws['G53'].value or 0
h53 = presentation_ws['H53'].value or 0
try:
    g53_num = float(g53)
except (TypeError, ValueError):
    g53_num = 0
try:
    h53_num = float(h53)
except (TypeError, ValueError):
    h53_num = 0
presentation_ws['G53'].value = g53_num + h53_num

# Replace H53 with value from AJ35 in financial report
aj35_value = financial_ws['AJ35'].value
presentation_ws['H53'].value = aj35_value

# --- Custom cell logic for "Presentation Working Sheet" In-Kind -----

# Sum F65 and E65 in Presentation and put result in E65
f65 = presentation_ws['F65'].value or 0
e65 = presentation_ws['E65'].value or 0
try:
    f65_num = float(f65)
except (TypeError, ValueError):
    f65_num = 0
try:
    e65_num = float(e65)
except (TypeError, ValueError):
    e65_num = 0
presentation_ws['E65'].value = f65_num + e65_num

# Replace F65 with value from AJ95 in financial report
aj95_value = financial_ws['AJ95'].value
presentation_ws['F65'].value = aj95_value


# --- Custom cell logic for "Presentation Working Sheet" MDA Contract Status -----
#pull data from acutals tab
#SKIP


#Summary Tab
# Copy AJ33, AJ35, AJ36 from financial report to I87, I88, I89 in Presentation Working Sheet
aj33_value = financial_ws['AJ33'].value
aj36_value = financial_ws['AJ36'].value

presentation_ws['I87'].value = aj33_value
presentation_ws['I88'].value = aj35_value
presentation_ws['I89'].value = aj36_value


target_wb.save('ATAC Project Dashboard Trial March.xlsx')