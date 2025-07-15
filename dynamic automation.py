from openpyxl import Workbook, load_workbook
from copy import copy
from datetime import datetime
import os 


def find_column_by_month(ws, month_label, header_row=3):
    """
    Find the column index for a given month label in the header row.
    month_label: e.g. "May-25" or "2025-05-01"
    header_row: the row number where months are listed (3 in your screenshot)
    """
    # Try to parse month_label as "May-25"
    try:
        target_date = datetime.strptime(month_label, "%b-%y")
    except ValueError:
        try:
            target_date = datetime.strptime(month_label, "%Y-%m-%d")
        except Exception:
            target_date = None

    for cell in ws[header_row]:
        if cell.value:
            # If cell is a date
            if isinstance(cell.value, datetime):
                if target_date and cell.value.year == target_date.year and cell.value.month == target_date.month:
                    return cell.column
            # If cell is a string (fallback)
            elif month_label.lower() in str(cell.value).lower():
                return cell.column
    return None

def find_section_total_row(ws, section_label, total_label="TOTAL", label_col=1):
    """Find the first 'TOTAL' row after a section label."""
    section_row = None
    for row in ws.iter_rows(min_col=label_col, max_col=label_col):
        cell = row[0]
        if cell.value and section_label.lower() in str(cell.value).lower():
            section_row = cell.row
            break
    if section_row:
        for row in ws.iter_rows(min_row=section_row+1, min_col=label_col, max_col=label_col):
            cell = row[0]
            if cell.value and total_label.lower() in str(cell.value).lower():
                return cell.row
    return None


from datetime import datetime

def find_column_by_header(ws, header_label, header_row=1):
    """Find the column index for a given header, handling both strings and Excel dates."""
    # Try to parse header_label as "May-25"
    try:
        target_date = datetime.strptime(header_label, "%b-%y")
    except Exception:
        target_date = None

    for cell in ws[header_row]:
        if cell.value:
            # If cell is a date
            if isinstance(cell.value, datetime) and target_date:
                if cell.value.year == target_date.year and cell.value.month == target_date.month:
                    return cell.column
            # If cell is a string (fallback)
            elif header_label.lower() in str(cell.value).lower():
                return cell.column
    return None


def find_first_row_by_label(ws, label="TOTAL", label_col=1):
    """Find the first row index where the cell value is exactly 'TOTAL' (all caps, no extra spaces)."""
    for row in ws.iter_rows(min_col=label_col, max_col=label_col):
        cell = row[0]
        if cell.value and str(cell.value).strip() == label:
            return cell.row
    return None

def find_row_by_label(ws, label, label_col=1):
    """Find the row index for a given label (case-insensitive, trimmed) in the specified column."""
    for row in ws.iter_rows(min_col=label_col, max_col=label_col):
        cell = row[0]
        if cell.value and label.lower() == str(cell.value).strip().lower():
            return cell.row
    return None

def find_row_by_label_after(ws, label, after_row, label_col=1):
    """Find the row index for a given label after a specific row."""
    for row in ws.iter_rows(min_row=after_row+1, min_col=label_col, max_col=label_col):
        cell = row[0]
        if cell.value and label.lower() == str(cell.value).strip().lower():
            return cell.row
    return None
# --- Main Script ---
# Financial Source Workbook
#financial_wb = load_workbook(r'C:\Users\Dsouzaj\Downloads\Automation excels\April\ATAC NRC Financial Report Apr 2025.xlsx', data_only=True)
financial_wb = load_workbook(r'C:\Users\Dsouzaj\Downloads\Automation excels\May\ATAC NRC Financial Report May 2025 (1).xlsx', data_only=True)

financial_ws = financial_wb['USD Monthly Totals']

# Dashboard Template Workbook


target_wb = load_workbook(r'C:\Users\Dsouzaj\Downloads\Automation excels\April\ATAC Project Dashbaord_May30, 25 with updated projections (1).xlsx')


presentation_ws = target_wb['Presentation Working Sheet']

# --- Custom cell logic for "Presentation Working Sheet" ALL PHASES TABLE-----

target_month = "May-25"



section_label = "ATAC Total Billable"

# Find the column for May-25 in the financial report
month_col = find_column_by_month(financial_ws, target_month, header_row=3)
print("month_col:", month_col)

# Find the TOTAL row under the section in the financial report
total_row = find_first_row_by_label(financial_ws, label="TOTAL", label_col=1)
print("total_row:", total_row)

if month_col and total_row:
    section_total = financial_ws.cell(row=total_row, column=month_col).value
    print(f"{section_label} TOTAL for {target_month}:", section_total)
    
# Find columns by header (header is in row 5 in your screenshot)
prev_inv_col = find_column_by_header(presentation_ws, "Previously Invoiced", header_row=5)
curr_inv_col = find_column_by_header(presentation_ws, "Current Invoice", header_row=5)

# Find the row for 2025 (year is in column B, which is col=2)
year_row = None
for row in presentation_ws.iter_rows(min_row=6, max_row=20, min_col=2, max_col=2):
    cell = row[0]
    if cell.value and str(cell.value).strip() == "2025":
        year_row = cell.row
        break

if prev_inv_col and curr_inv_col and year_row:
    # Step 1: Get the original values
    prev_val = presentation_ws.cell(row=year_row, column=prev_inv_col).value or 0
    curr_val = presentation_ws.cell(row=year_row, column=curr_inv_col).value or 0
    total = prev_val + curr_val
    print(f"Table #1: Sum of Previously Invoiced and Current Invoice for 2025: {total}")

    # (Optional) Step 2: Update Previously Invoiced (G9) with the sum
    presentation_ws.cell(row=year_row, column=prev_inv_col).value = total

    # Step 3: Replace Current Invoice (H9) with the value from May-25
    presentation_ws.cell(row=year_row, column=curr_inv_col).value = section_total
    print(f"Table #1: Replaced Current Invoice for 2025 with {section_total}")
else:
    print("Could not find required columns or row for 2025.")

# --- Second Table: Revised table - Labour + Travel ---

labour_table_header_row = 18  # Use the row with the actual column headers
labour_table_first_data_row = 20  # First year row (2023), so 2025 is at row 22

prev_inv_col2 = find_column_by_header(presentation_ws, "Previously Invoiced", header_row=labour_table_header_row)
curr_inv_col2 = find_column_by_header(presentation_ws, "Current Invoice", header_row=labour_table_header_row)


# Find the row for 2025 in the second table (column B)
year_row2 = None
for row in presentation_ws.iter_rows(min_row=labour_table_first_data_row, max_row=labour_table_first_data_row+10, min_col=2, max_col=2):
    cell = row[0]
    if cell.value and str(cell.value).strip() == "2025":
        year_row2 = cell.row
        break
    # Handle if cell.value is int 2025
    if cell.value and isinstance(cell.value, int) and cell.value == 2025:
        year_row2 = cell.row
        break
print("prev_inv_col2:", prev_inv_col2)
print("curr_inv_col2:", curr_inv_col2)
print("year_row2:", year_row2)


# --- Get Labour + Travel from Financial Report ---

def get_numeric(ws, row, col):
    """Return the numeric value or 0 if None, dash, or not a number."""
    val = ws.cell(row=row, column=col).value
    if val is None or (isinstance(val, str) and val.strip() in ["", "-"]):
        return 0
    try:
        return float(val)
    except Exception:
        return 0

# Hardcode the row numbers for Labour and Travel based on your screenshot
labour_row = find_row_by_label(financial_ws, "Labour", label_col=1)
travel_row = find_row_by_label(financial_ws, "Travel", label_col=1)

if month_col and curr_inv_col2 and year_row2 and prev_inv_col2:
    # Get original values BEFORE replacing
    prev_val2 = presentation_ws.cell(row=year_row2, column=prev_inv_col2).value or 0
    curr_val2 = presentation_ws.cell(row=year_row2, column=curr_inv_col2).value or 0
    total2 = prev_val2 + curr_val2

    # Step 1: Update Previously Invoiced with the sum
    presentation_ws.cell(row=year_row2, column=prev_inv_col2).value = total2
    print(f"Table #2: Updated Labour+Travel Previously Invoiced for 2025: {total2}")

    # Step 2: Update Current Invoice for 2025 in the dashboard's second table
    labour_val = get_numeric(financial_ws, labour_row, month_col)
    travel_val = get_numeric(financial_ws, travel_row, month_col)
    sum_labour_travel = labour_val + travel_val
    presentation_ws.cell(row=year_row2, column=curr_inv_col2).value = sum_labour_travel
    print(f"Table #2: Updated Labour+Travel Current Invoice for 2025 with {sum_labour_travel}")
else:
    print("Could not find required columns, rows, or month for Labour+Travel table.")

#Table #3: Labour and Travel Actual
# Find the column for May-25 in the financial report and dashboard

fin_month_col = find_column_by_month(financial_ws, target_month, header_row=3)
dash_month_col = find_column_by_header(presentation_ws, target_month, header_row=39)  # Adjust header_row if needed


# Find the NRC and MDA rows in the dashboard (assuming column B has the labels)
nrc_dash_row = find_row_by_label(presentation_ws, "NRC", label_col=2)
mda_dash_row = find_row_by_label(presentation_ws, "MDA", label_col=2)


# NRC
# Find the NRC Billable Labour Sub Total row in the financial report
nrc_total_row = find_row_by_label(financial_ws, "Total NRC Billable Labour Sub Total", label_col=1)
if fin_month_col and nrc_total_row and dash_month_col and nrc_dash_row:
    raw_nrc_val = financial_ws.cell(row=nrc_total_row, column=fin_month_col).value
    nrc_val = get_numeric(financial_ws, nrc_total_row, fin_month_col)
    nrc_val_rounded = round(nrc_val)  # Round to nearest integer
    nrc_cell = presentation_ws.cell(row=nrc_dash_row, column=dash_month_col)
    nrc_cell.value = nrc_val_rounded
    # Copy style from previous cell
    ref_cell = presentation_ws.cell(row=nrc_dash_row, column=dash_month_col - 1)
    nrc_cell.font = copy(ref_cell.font)
    nrc_cell.number_format = copy(ref_cell.number_format)
    nrc_cell.alignment = copy(ref_cell.alignment)
    print(f"Copied NRC value {nrc_val_rounded} to dashboard at row {nrc_dash_row}, col {dash_month_col}.")

# MDA

subcon_labour_row = find_row_by_label(financial_ws, "Sub-Contractor Labour", label_col=1)

if fin_month_col and subcon_labour_row and dash_month_col and mda_dash_row:
    raw_subcon_labour_val = financial_ws.cell(row=subcon_labour_row, column=fin_month_col).value
    subcon_labour_val = get_numeric(financial_ws, subcon_labour_row, fin_month_col)
    subcon_labour_val_rounded = round(subcon_labour_val)
    mda_cell = presentation_ws.cell(row=mda_dash_row, column=dash_month_col)
    mda_cell.value = subcon_labour_val_rounded
    # Copy style from previous cell
    ref_cell = presentation_ws.cell(row=mda_dash_row, column=dash_month_col - 1)
    mda_cell.font = copy(ref_cell.font)
    mda_cell.number_format = copy(ref_cell.number_format)
    mda_cell.alignment = copy(ref_cell.alignment)
    print(f"Copied Sub-Contractor Labour value {subcon_labour_val_rounded} to dashboard at row {mda_dash_row}, col {dash_month_col}.")


# Find the "Total (USD)" row in the dashboard (assuming column B has the label)
actual_row = find_row_by_label(presentation_ws, "Actual", label_col=2)
total_dash_row = find_row_by_label_after(presentation_ws, "Total (USD)", after_row=actual_row, label_col=2)

if dash_month_col and nrc_dash_row and mda_dash_row and total_dash_row:
    nrc_val = presentation_ws.cell(row=nrc_dash_row, column=dash_month_col).value or 0
    mda_val = presentation_ws.cell(row=mda_dash_row, column=dash_month_col).value or 0
    total_val = nrc_val + mda_val
    total_cell = presentation_ws.cell(row=total_dash_row, column=dash_month_col)
    total_cell.value = total_val
    # Copy style from previous cell
    ref_cell = presentation_ws.cell(row=total_dash_row, column=dash_month_col - 1)
    total_cell.font = copy(ref_cell.font)

    total_cell.number_format = copy(ref_cell.number_format)
    total_cell.alignment = copy(ref_cell.alignment)
    print("Total value:", total_val)
    print(f"Copied Total (USD) value {total_val} to dashboard at row {total_dash_row}, col {dash_month_col}.")


# SKIPPED %COMPLETE  ROW

# --- Revised Table (NRE + SI) ---

# Adjust these row numbers if your table layout changes
nre_table_header_row = 59  # The row with the column headers for this table
nre_table_first_data_row = 61  # The row where 2025 appears in this table

# Find columns for "Previously Invoiced" and "Current Invoice"
prev_inv_col3 = find_column_by_header(presentation_ws, "Previously Invoiced", header_row=nre_table_header_row)
curr_inv_col3 = find_column_by_header(presentation_ws, "Current Invoice", header_row=nre_table_header_row)

# Find the row for 2025 in this table (column B)
year_row3 = None
for row in presentation_ws.iter_rows(min_row=nre_table_first_data_row, max_row=nre_table_first_data_row+10, min_col=2, max_col=2):
    cell = row[0]
    if cell.value and str(cell.value).strip() == "2025":
        year_row3 = cell.row
        break
    if cell.value and isinstance(cell.value, int) and cell.value == 2025:
        year_row3 = cell.row
        break

if prev_inv_col3 and curr_inv_col3 and year_row3:
    # Step 1: Add current invoice to previously invoiced and update
    prev_val3 = presentation_ws.cell(row=year_row3, column=prev_inv_col3).value or 0
    curr_val3 = presentation_ws.cell(row=year_row3, column=curr_inv_col3).value or 0
    try:
        prev_val3 = float(str(prev_val3).replace(",", ""))
    except Exception:
        prev_val3 = 0
    try:
        curr_val3 = float(str(curr_val3).replace(",", ""))
    except Exception:
        curr_val3 = 0
    total3 = prev_val3 + curr_val3
    presentation_ws.cell(row=year_row3, column=prev_inv_col3).value = total3
    print(f"Revised Table (NRE + SI): Updated Previously Invoiced for 2025: {total3}")

    # Step 2: Get NRE & SI value from financial report for the target month
    nre_row = find_row_by_label(financial_ws, "NRE & SI", label_col=1)
    nre_val = get_numeric(financial_ws, nre_row, month_col) if nre_row and month_col else 0
    presentation_ws.cell(row=year_row3, column=curr_inv_col3).value = nre_val
    print(f"Revised Table (NRE + SI): Updated Current Invoice for 2025 with {nre_val}")
else:
    print("Could not find required columns or row for Revised Table (NRE + SI).")


# -------------IN-KIND TABLE--------------

# --- In-Kind Table Update ---

# Dashboard: Find columns for "Previously Amount" and "Current Amount" in the In-Kind table
inkind_table_header_row = 71  # Row with headers in dashboard (adjust if needed)
inkind_table_first_data_row = 73  # First year row (2023), so 2025 is at row 75

prev_amt_col = find_column_by_header(presentation_ws, "Previously Amount", header_row=inkind_table_header_row)
curr_amt_col = find_column_by_header(presentation_ws, "Current Amount", header_row=inkind_table_header_row)

# Find the row for 2025 in the In-Kind table (column B)
inkind_2025_row = None
for row in presentation_ws.iter_rows(min_row=inkind_table_first_data_row, max_row=inkind_table_first_data_row+10, min_col=2, max_col=2):
    cell = row[0]
    if cell.value and str(cell.value).strip() == "2025":
        inkind_2025_row = cell.row
        break
    if cell.value and isinstance(cell.value, int) and cell.value == 2025:
        inkind_2025_row = cell.row
        break

if prev_amt_col and curr_amt_col and inkind_2025_row:
    # Step 1: Add current amount to previous amount and update
    prev_val = presentation_ws.cell(row=inkind_2025_row, column=prev_amt_col).value or 0
    curr_val = presentation_ws.cell(row=inkind_2025_row, column=curr_amt_col).value or 0
    try:
        prev_val = float(str(prev_val).replace(",", ""))
    except Exception:
        prev_val = 0
    try:
        curr_val = float(str(curr_val).replace(",", ""))
    except Exception:
        curr_val = 0
    total_val = prev_val + curr_val
    presentation_ws.cell(row=inkind_2025_row, column=prev_amt_col).value = total_val
    print(f"In-Kind Table: Updated Previously Amount for 2025: {total_val}")

# 1. Find section row
section_row = None
for row in financial_ws.iter_rows(min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "atac total in-kind" in str(cell.value).lower():
        section_row = cell.row
        print(f"Found section at row {section_row}: {cell.value}")
        break

# 2. Find TOTAL row
total_row = None
if section_row:
    for row in financial_ws.iter_rows(min_row=section_row+1, min_col=1, max_col=1):
        cell = row[0]
        print(f"Checking row {cell.row}: {repr(cell.value)}")
        if cell.value and str(cell.value).strip().upper() == "TOTAL":
            total_row = cell.row
            print(f"Found TOTAL at row {total_row}")
            break

# 3. Find May-25 column
month_col = find_column_by_month(financial_ws, "May-25", header_row=85)  # Adjust header_row if needed
print(f"May-25 column: {month_col}")

# 4. Get value and paste to dashboard
if total_row and month_col and inkind_2025_row and curr_amt_col:
    total_val = financial_ws.cell(row=total_row, column=month_col).value
    print(f"TOTAL value for May-25: {total_val}")
    presentation_ws.cell(row=inkind_2025_row, column=curr_amt_col).value = total_val
    print(f"Pasted {total_val} to dashboard row {inkind_2025_row}, col {curr_amt_col}")
else:
    print("Could not find TOTAL row, May-25 column, inkind_2025_row, or curr_amt_col.")


#----------SKIP CALL UP TABLE-----------

#-----------SUMMARY TABLE-------------

# Find month column in financial report and dashboard
month_col = find_column_by_month(financial_ws, "May-25", header_row=3)
dash_month_col = find_column_by_header(presentation_ws, "May-25", header_row= 104 )  # Set this

# Find rows in financial report
labour_row = find_row_by_label(financial_ws, "Labour", label_col=1)
travel_row = find_row_by_label(financial_ws, "Travel", label_col=1)
nre_row = find_row_by_label(financial_ws, "NRE & SI", label_col=1)
bom_row = find_row_by_label(financial_ws, "BOM", label_col=1)

# Get values
labour_val = get_numeric(financial_ws, labour_row, month_col)
travel_val = get_numeric(financial_ws, travel_row, month_col)
nre_val = get_numeric(financial_ws, nre_row, month_col)
sum_labour_travel = labour_val + travel_val
bom_val = get_numeric(financial_ws,bom_row, month_col)

# Find dashboard rows
labour_travel_dash_row = find_row_by_label(presentation_ws, "Labour+Travel", label_col=2)
nre_dash_row = find_row_by_label(presentation_ws, "NRE + SI", label_col=2)
bom_dash_row = find_row_by_label(presentation_ws, "BOM", label_col=2)

# Paste values
if labour_travel_dash_row and dash_month_col:
    presentation_ws.cell(row=labour_travel_dash_row, column=dash_month_col).value = round(sum_labour_travel)
    print(f"Pasted Labour+Travel {round(sum_labour_travel)} to dashboard at row {labour_travel_dash_row}, col {dash_month_col}")


if nre_dash_row and dash_month_col:
    presentation_ws.cell(row=nre_dash_row, column=dash_month_col).value = round(nre_val)
    print(f"Pasted NRE & SI {nre_val} to dashboard at row {nre_dash_row}, col {dash_month_col}")

if bom_dash_row and dash_month_col:
    presentation_ws.cell(row=bom_dash_row, column=dash_month_col).value = round(bom_val)
    print(f"Pasted BOM {bom_val} to dashboard at row {bom_dash_row}, col {dash_month_col}")


target_wb.save(r"C:\Users\Dsouzaj\Documents\NRC Python Files\Project Dashboard Automation Repo\ATAC Project Dashboard Trial May.xlsx")







